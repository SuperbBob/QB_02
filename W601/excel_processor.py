"""
Excel File Processor Module
Handles complex Excel table reshaping into 2D tables
Based on reference: dismantle_excel.py
"""
import json
import logging
import os
import uuid
import hashlib
from pathlib import Path
from typing import Dict, List, Optional, Any
from collections import OrderedDict

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(f'excel_agent.{__name__}')


class ExcelProcessor:
    """Process and reshape complex Excel files into standardized 2D tables"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.file_name = self.file_path.name
        self.sheets: Dict[str, pd.DataFrame] = {}
        self.metadata: Dict[str, Any] = {}
        self.merged_info: Dict[str, List] = {}
        
    def load_file(self) -> bool:
        """Load Excel file and all its sheets"""
        try:
            if self.file_path.suffix.lower() == '.csv':
                df = pd.read_csv(self.file_path)
                self.sheets['Sheet1'] = df
            else:
                xlsx = pd.ExcelFile(self.file_path)
                for sheet_name in xlsx.sheet_names:
                    self.sheets[sheet_name] = pd.read_excel(
                        xlsx, sheet_name=sheet_name, header=None
                    )
            return True
        except Exception as e:
            logger.error(f"Error loading file {self.file_path}: {e}")
            return False
    
    def unmerge_and_fill_cells(self, output_path: str) -> tuple:
        """
        Unmerge all merged cells and fill with original values
        Returns: (output_path, merged_info dict)
        """
        try:
            logger.info("Starting to unmerge all cells...")
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            merged_info = {}
            
            for ws in wb.worksheets:
                logger.info(f"Processing sheet: {ws.title}")
                sheet_merged_info = []
                
                for merged_range in list(ws.merged_cells.ranges):
                    min_row, min_col = merged_range.min_row, merged_range.min_col
                    max_row, max_col = merged_range.max_row, merged_range.max_col
                    value = ws.cell(row=min_row, column=min_col).value
                    
                    # Collect merged cell info for first 6 rows (header area)
                    if max_row <= 6:
                        sheet_merged_info.append({
                            "merged_range": str(merged_range),
                            "start_cell": (min_row, min_col),
                            "end_cell": (max_row, max_col),
                            "value": value
                        })
                    
                    # Unmerge cells
                    ws.unmerge_cells(
                        start_row=min_row, start_column=min_col,
                        end_row=max_row, end_column=max_col
                    )
                    
                    # Fill all split cells with the original value
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            ws.cell(row=row, column=col, value=value)
                
                merged_info[ws.title] = sheet_merged_info
            
            wb.save(output_path)
            self.merged_info = merged_info
            logger.info(f"Unmerged file saved to: {output_path}")
            return output_path, merged_info
            
        except Exception as e:
            logger.error(f"Error unmerging cells: {e}", exc_info=True)
            return None, {}
    
    def get_excel_preview(self, head: int = 6) -> List[str]:
        """Get preview of first N rows for each sheet"""
        try:
            all_sheets_data = pd.read_excel(
                self.file_path, sheet_name=None, header=None
            )
            prompt_parts = []
            
            for sheet_name, data in all_sheets_data.items():
                data.index = data.index + 1
                excel_col_names = [get_column_letter(i + 1) for i in range(len(data.columns))]
                data.columns = excel_col_names
                
                # Handle newlines in cells
                data = data.map(
                    lambda x: str(x).replace('\n', ' ') if isinstance(x, str) else x
                )
                
                sheet_first_rows = data.head(head).to_markdown(index=True)
                sheet_info = f"Sheet: {sheet_name}\n前 {head} 行数据为：\n\n{sheet_first_rows}\n\n---"
                prompt_parts.append(sheet_info)
                
            return prompt_parts
        except Exception as e:
            logger.error(f"Error getting preview: {e}", exc_info=True)
            return []
    
    def detect_header_row(self, df: pd.DataFrame) -> int:
        """Detect the actual header row in a DataFrame"""
        max_non_null_row = 0
        max_non_null_count = 0
        
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            non_null_count = row.notna().sum()
            string_count = sum(1 for val in row if isinstance(val, str) and val.strip())
            
            if string_count > max_non_null_count:
                max_non_null_count = string_count
                max_non_null_row = idx
                
        return max_non_null_row
    
    def deduplication_header(
        self, 
        df: pd.DataFrame, 
        header_rows: List[int]
    ) -> pd.DataFrame:
        """
        Deduplicate and merge multi-level headers
        """
        try:
            if len(header_rows) == 1:
                # Single header row
                df.columns = df.iloc[header_rows[0] - 1].fillna('').astype(str)
                df = df.iloc[header_rows[0]:].reset_index(drop=True)
            else:
                # Multi-level headers - merge them
                header_data = [df.iloc[row - 1].fillna('').astype(str).tolist() 
                              for row in header_rows]
                
                new_columns = []
                for col_idx in range(len(header_data[0])):
                    col_parts = [header_data[row][col_idx] for row in range(len(header_data))]
                    # Deduplicate while preserving order
                    dedup_parts = list(OrderedDict.fromkeys(col_parts))
                    valid_header = '-'.join([
                        part for part in dedup_parts 
                        if part and 'Unnamed' not in part and part != 'nan'
                    ])
                    new_columns.append(valid_header if valid_header else f'Column_{col_idx}')
                
                df.columns = new_columns
                df = df.iloc[max(header_rows):].reset_index(drop=True)
            
            return df
        except Exception as e:
            logger.error(f"Error deduplicating headers: {e}", exc_info=True)
            return df
    
    def reshape_to_2d(
        self, 
        df: pd.DataFrame,
        labels_to_drop: List[int] = None,
        header_rows: List[int] = None
    ) -> pd.DataFrame:
        """
        Reshape complex Excel tables into clean 2D format
        Handles merged cells, multi-level headers, etc.
        """
        # Drop label/description rows if specified
        if labels_to_drop:
            # Convert to 0-based index
            labels_idx = [x - 1 for x in labels_to_drop if x - 1 < len(df)]
            df = df.drop(labels_idx, axis=0, errors='ignore').reset_index(drop=True)
        
        # Handle empty DataFrame
        if df.empty or len(df) == 0:
            return df
        
        # Handle headers
        if header_rows:
            # Adjust header rows after dropping labels
            adjusted_headers = [
                x - len([l for l in (labels_to_drop or []) if l < x]) 
                for x in header_rows
            ]
            df = self.deduplication_header(df.copy(), adjusted_headers)
        else:
            # Auto-detect header row
            header_row = self.detect_header_row(df)
            if header_row > 0 and header_row < len(df):
                df = df.iloc[header_row:].reset_index(drop=True)
            
            # Check if we still have data after header detection
            if len(df) > 0:
                df.columns = df.iloc[0].fillna('').astype(str)
                if len(df) > 1:
                    df = df.iloc[1:].reset_index(drop=True)
                else:
                    # Only header row, no data
                    df = df.iloc[0:0]  # Empty with column names
        
        # Return if empty after processing
        if df.empty:
            return df
        
        # Clean column names
        df.columns = [self._clean_column_name(col) for col in df.columns]
        df = self._handle_duplicate_columns(df)
        
        # Remove completely empty rows and columns
        df = df.dropna(how='all', axis=0)
        df = df.dropna(how='all', axis=1)
        
        # Convert data types
        df = self._infer_and_convert_types(df)
        
        return df
    
    def _clean_column_name(self, name: str) -> str:
        """Clean and standardize column names"""
        if not name or str(name).strip() == '' or str(name) == 'nan':
            return 'Unnamed'
        name = str(name).strip()
        name = ' '.join(name.split())
        return name
    
    def _handle_duplicate_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Handle duplicate column names by adding suffixes"""
        cols = df.columns.tolist()
        seen = {}
        new_cols = []
        
        for col in cols:
            if col in seen:
                seen[col] += 1
                new_cols.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                new_cols.append(col)
                
        df.columns = new_cols
        return df
    
    def _infer_and_convert_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """Infer and convert column data types with proper handling"""
        import warnings
        warnings.simplefilter(action='ignore', category=Warning)
        
        for col in df.columns:
            # Try numeric conversion
            try:
                numeric_series = pd.to_numeric(df[col], errors='coerce')
                if numeric_series.notna().sum() / max(len(df), 1) > 0.5:
                    df[col] = numeric_series
                    continue
            except:
                pass
            
            # Try datetime conversion
            try:
                datetime_series = pd.to_datetime(df[col], errors='coerce')
                if datetime_series.notna().sum() / max(len(df), 1) > 0.5:
                    df[col] = datetime_series.dt.normalize()
                    continue
            except:
                pass
                
        return df
    
    def process_all_sheets(
        self, 
        sheet_config: Dict[str, Dict] = None
    ) -> Dict[str, pd.DataFrame]:
        """
        Process all sheets and return cleaned DataFrames
        sheet_config format: {
            "sheet_name": {
                "labels": [row_numbers_to_drop],
                "header": [header_row_numbers]
            }
        }
        """
        processed = {}
        for sheet_name, df in self.sheets.items():
            config = (sheet_config or {}).get(sheet_name, {})
            labels = config.get('labels', [])
            headers = config.get('header', None)
            
            processed[sheet_name] = self.reshape_to_2d(
                df.copy(),
                labels_to_drop=labels,
                header_rows=headers
            )
        return processed
    
    def get_file_summary(self) -> Dict[str, Any]:
        """Generate a summary of the Excel file structure"""
        summary = {
            "file_name": self.file_name,
            "file_path": str(self.file_path),
            "sheets": []
        }
        
        for sheet_name, df in self.sheets.items():
            try:
                processed_df = self.reshape_to_2d(df.copy())
            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name}: {e}")
                processed_df = df.copy()
            
            columns_info = []
            if not processed_df.empty:
                for col in processed_df.columns:
                    try:
                        col_info = {
                            "name": str(col),
                            "dtype": str(processed_df[col].dtype),
                            "non_null_count": int(processed_df[col].notna().sum()),
                            "sample_values": [
                                str(v) for v in processed_df[col].dropna().head(3).tolist()
                            ]
                        }
                        columns_info.append(col_info)
                    except Exception:
                        columns_info.append({"name": str(col), "dtype": "unknown", "non_null_count": 0, "sample_values": []})
            
            sheet_info = {
                "sheet_name": sheet_name,
                "row_count": len(processed_df),
                "column_count": len(processed_df.columns),
                "columns": columns_info
            }
            summary["sheets"].append(sheet_info)
            
        return summary
    
    def get_column_descriptions(self) -> str:
        """Generate natural language descriptions of columns for LLM context"""
        descriptions = []
        
        for sheet_name, df in self.sheets.items():
            processed_df = self.reshape_to_2d(df.copy())
            sheet_desc = f"\n【工作表: {sheet_name}】\n"
            sheet_desc += f"行数: {len(processed_df)}, 列数: {len(processed_df.columns)}\n"
            sheet_desc += "列信息:\n"
            
            for col in processed_df.columns:
                dtype = processed_df[col].dtype
                non_null = processed_df[col].notna().sum()
                samples = processed_df[col].dropna().head(3).tolist()
                
                sheet_desc += f"  - {col}: 类型={dtype}, 非空数={non_null}, 示例值={samples}\n"
                
            descriptions.append(sheet_desc)
            
        return "\n".join(descriptions)


class KnowledgeBase:
    """Manage a collection of Excel files as a knowledge base"""
    
    def __init__(self, base_path: str):
        self.base_path = Path(base_path)
        self.base_path.mkdir(exist_ok=True)
        self.files: Dict[str, ExcelProcessor] = {}
        self.index: Dict[str, Any] = {}
        
    def add_file(self, file_path: str) -> bool:
        """Add a file to the knowledge base"""
        processor = ExcelProcessor(file_path)
        if processor.load_file():
            file_id = self._generate_file_id(file_path)
            self.files[file_id] = processor
            self.index[file_id] = processor.get_file_summary()
            return True
        return False
    
    def scan_directory(self) -> int:
        """Scan directory for Excel files and add them to knowledge base"""
        count = 0
        for ext in ['.xlsx', '.xls', '.csv']:
            for file_path in self.base_path.glob(f'*{ext}'):
                if self.add_file(str(file_path)):
                    count += 1
        return count
    
    def _generate_file_id(self, file_path: str) -> str:
        """Generate unique ID for a file"""
        return hashlib.md5(file_path.encode()).hexdigest()[:12]
    
    def get_all_summaries(self) -> str:
        """Get combined summaries of all files for LLM context"""
        summaries = []
        for file_id, summary in self.index.items():
            summaries.append(f"\n=== 文件: {summary['file_name']} (ID: {file_id}) ===")
            for sheet in summary['sheets']:
                summaries.append(f"\n工作表: {sheet['sheet_name']}")
                summaries.append(f"数据量: {sheet['row_count']}行 x {sheet['column_count']}列")
                summaries.append("列信息:")
                for col in sheet['columns']:
                    samples = ', '.join(col.get('sample_values', [])[:2])
                    summaries.append(f"  - {col['name']} ({col['dtype']}): 示例=[{samples}]")
        return "\n".join(summaries)
    
    def find_relevant_files(self, query: str, columns_mentioned: List[str]) -> List[str]:
        """Find files that contain relevant columns for the query"""
        relevant_files = []
        
        for file_id, summary in self.index.items():
            for sheet in summary['sheets']:
                column_names = [col['name'].lower() for col in sheet['columns']]
                
                for col in columns_mentioned:
                    if col.lower() in column_names or any(col.lower() in c for c in column_names):
                        relevant_files.append(file_id)
                        break
                        
        return list(set(relevant_files))
    
    def get_file_by_id(self, file_id: str) -> Optional[ExcelProcessor]:
        """Get file processor by ID"""
        return self.files.get(file_id)
    
    def get_all_files(self) -> Dict[str, ExcelProcessor]:
        """Get all files in knowledge base"""
        return self.files
